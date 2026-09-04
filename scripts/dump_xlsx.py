# -*- coding: utf-8 -*-
"""只读 dump xlsx 全部内容为 JSON，含合并单元格/超链接/填充色/批注信息，不做任何修改。"""
import json, sys, io
from openpyxl import load_workbook

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

PATH = r"C:\Users\Lenovo\Desktop\秋招简历\27秋招信息整理.xlsx"
wb = load_workbook(PATH, data_only=True, read_only=False)

out = {"file": PATH, "sheets": []}
for ws in wb.worksheets:
    sheet = {
        "name": ws.title,
        "max_row": ws.max_row,
        "max_col": ws.max_column,
        "merged": [str(r) for r in ws.merged_cells.ranges],
        "hyperlinks": {},
        "fills": {},          # 非默认填充色: A1 -> hex
        "font_colors": {},    # 非默认字体色: A1 -> hex
        "rows": [],
    }
    # 表头行假定第 1 行；若第1行空则取第一个有内容的行作为表头
    header_row = 1
    for r in range(1, min(ws.max_row, 5) + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if any(v is not None and str(v).strip() != "" for v in vals):
            header_row = r
            break
    sheet["header_row"] = header_row

    for row in ws.iter_rows():
        r_idx = row[0].row
        vals = []
        for cell in row:
            vals.append(cell.value)
        sheet["rows"].append(vals)

    for h in ws._hyperlinks:
        sheet["hyperlinks"][h.ref] = h.target if h.target else h.location
    # 填充/字体颜色（仅记非默认）
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            fill = cell.fill
            fg = None
            try:
                if fill and fill.patternType and fill.fgColor and fill.fgColor.rgb and str(fill.fgColor.rgb) not in ("00000000",):
                    fg = str(fill.fgColor.rgb)
            except Exception:
                pass
            if fg:
                sheet["fills"][cell.coordinate] = fg
            fc = None
            try:
                if cell.font and cell.font.color and cell.font.color.rgb and str(cell.font.color.rgb) not in ("00000000", "FF000000"):
                    fc = str(cell.font.color.rgb)
            except Exception:
                pass
            if fc:
                sheet["font_colors"][cell.coordinate] = fc
    out["sheets"].append(sheet)

print(json.dumps(out, ensure_ascii=False, default=str))
